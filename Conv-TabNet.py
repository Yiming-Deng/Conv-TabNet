import numpy as np
import os
from pandas.core.frame import DataFrame
import pandas as pd
from visdom import Visdom
import math
import random

from matplotlib import pyplot as plt
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import normalize
from sklearn.metrics import mean_squared_error, mean_absolute_error
from sklearn.model_selection import train_test_split

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors

from pytorch_tabnet.tab_model import TabNetRegressor

from sklearn.metrics import r2_score
from tqdm import tqdm


def ColourDistance(rgb_1, rgb_2):
    R_1, G_1, B_1 = rgb_1
    R_2, G_2, B_2 = rgb_2
    rmean = (R_1 + R_2) / 2
    R = R_1 - R_2
    G = G_1 - G_2
    B = B_1 - B_2
    return math.sqrt((2 + rmean / 256) * (R**2) + 4 * (G**2) + (2 + (255 - rmean) / 256) * (B**2))


def RGB_to_Hex(rgb):
    strs = '#'
    for i in rgb:
        num = int(float(i))  # 将str转int
        # 将R、G、B分别转化为16进制拼接转换并大写
        strs += str(hex(num))[-2:].replace('x', '0').upper()

    return strs


def arr2str(arr):
    # data = []
    hex_list = []
    for i in arr:
        # s = ','.join(str(j) for j in i)
        hex_list.append(RGB_to_Hex(i))
        # data.append(s)
    # data = np.array(data).reshape(-1, 1)
    hex_list = np.array(hex_list).reshape(-1, 1)
    return np.hstack((arr, hex_list))


def rmse(pred, label):
    tot = 0
    for i in range(len(pred)):
        rmse = 0
        for j in range(len(pred[i])):
            rmse += (pred[i][j] - label[i][j])**2
        rmse = np.sqrt(rmse / (len(pred[i])))
        tot += rmse

    return tot / (len(pred))


def rae(pred, label):
    rae = 0
    label_aver = 0
    for i in range(len(pred)):
        up = 0
        down = 0
        for j in range(len(pred[i])):
            up += abs(pred[i][j] - label[i][j])
            down += label[i][j]
        rae += up / down
        label_aver += down

    return rae / len(pred) * 100


if __name__ == '__main__':
    # 输入数据
    # tot_path = r'D:\VScodeProjects\PythonProjects\ColorCorrection\Code\Dataset\20220721\result.xlsx'
    tot_path = r'D:\VScodeProjects\PythonProjects\ColorCorrection\Code\Dataset\tot_result.xlsx'
    # train_Path = r'D:\VScodeProjects\PythonProjects\ColorCorrection\Code\Dataset\train.xlsx'
    # test_Path = r'D:\VScodeProjects\PythonProjects\ColorCorrection\Code\Dataset\test.xlsx'
    # 输出表格
    result_path = r'D:\VScodeProjects\PythonProjects\ColorCorrection\Code\output_1124.xlsx'

    df = pd.read_excel(tot_path)
    df = df.loc[df['f_distance'] != 0.0]
    # df = df.loc[df['mobile'] != 'vivo']

    data = np.array(df)
    train_data, test_data = train_test_split(data, train_size=0.8, random_state=42, shuffle=True)

    # 载入输入数据
    # df = pd.read_excel(train_Path)
    # train_data = np.array(df)
    # train_input = train_data[:, 3:9]
    train_input = np.hstack((train_data[:, 3:9], train_data[:, 10:12], train_data[:, 12:14]))
    train_input = train_input.astype(np.float32)

    # 载入输出数据
    # train_label = np.array(pd.read_csv(train_Path))
    train_label = train_data[:, 15:18]
    train_label = train_label.astype(np.float32)

    # cc = list(zip(train_input, train_label))
    # random.shuffle(cc)
    # train_input[:], train_label[:] = zip(*cc)

    # 载入测试集
    # test_data = np.array(pd.read_excel(test_Path))
    # test_input = test_data[:, 3:9]
    test_input = np.hstack((test_data[:, 3:9], test_data[:, 10:12], test_data[:, 12:14]))
    test_input = test_input.astype(np.float32)

    # test_label = pd.read_csv(test_Path)
    test_label = test_data[:, 15:18]
    test_label = test_label.astype(np.float32)

    # valid = []
    # for i in tqdm(range(5)):
    #     valid.append((test_input[i].reshape(1, -1), test_label[i].reshape(1, -1)))

    # 网络
    net = TabNetRegressor()

    model_save_path = r'D:\VScodeProjects\PythonProjects\ColorCorrection\Code\Model_tab0113'
    # net.load_model(model_save_path + '.zip')

    # print(train_input[0])
    # input()
    # 模型训练
    # for i in range(100):
    feature_importances = net.fit(
        train_input,
        train_label,
        max_epochs=50000,
        batch_size=4096,
        #   eval_set=valid,
        #   eval_metric=['rmse'],
        patience=0)  # epoch：50000
    feature_dataframe = pd.DataFrame(feature_importances)
    feature_dataframe.to_csv('feature_importances_conv.csv')

    # 模型保存
    model_save_path = r'D:\VScodeProjects\PythonProjects\ColorCorrection\Code\Model_0106'
    # net.load_model(model_save_path + '.zip'
    saved_filepath = net.save_model(model_save_path)

    # 预测
    train_pred = net.predict(train_input)
    test_preds = net.predict(test_input)

    # train_pred = net.predict(train_input)
    # test_preds = net.predict(test_input)

    # 评价
    print('train myRMSE:{:.3f}'.format(rmse(train_pred, train_label)))
    print('train RMSE:{:.3f}'.format(np.sqrt(mean_squared_error(train_label, train_pred))))
    print('train MAE:{:.3f}'.format(mean_absolute_error(train_label, train_pred)))
    print('train R2:{:.3f}'.format(r2_score(train_pred, train_label)))
    print('train RAE:{:.3f}%'.format(rae(train_pred, train_label)))
    print('test myRMSE:{:.3f}'.format(rmse(test_preds, test_label)))
    print('test RMSE:{:.3f}'.format(np.sqrt(mean_squared_error(test_preds, test_label))))
    print('test MAE:{:.3F}'.format(mean_absolute_error(test_label, test_preds)))
    print('test R2:{:.3f}'.format(r2_score(test_preds, test_label)))
    print('data RAE:{:.3f}%'.format(rae(test_preds, test_label)))

    # 预测值保存
    # preds = np.clip(test_preds, 0, 255).astype(int)
    # preds = arr2str(preds)
    # Input = arr2str(test_input[:, :3].astype(int))
    # label = arr2str(test_label.astype(int))

    # # preds = np.clip(train_pred, 0, 255).astype(int)
    # # preds = arr2str(preds)
    # # Input = arr2str(train_input[:, :3].astype(int))
    # # label = arr2str(train_label.astype(int))
    # result = np.hstack((Input, preds, label))

    # dataframe = pd.DataFrame(result,
    #                          columns=[
    #                              'input_R', 'input_G', 'input_B', 'input_color', 'out_R', 'out_G', 'out_B', 'out_color',
    #                              'label_R', 'label_G', 'label_B', 'label_color'
    #                          ])
    # dataframe.to_excel(result_path, index=False)

    # wb = load_workbook(result_path)
    # sheet = wb.worksheets[0]

    # for i in [4, 8, 12]:
    #     for j in range(2, len(test_input) + 2):
    #         cell = sheet.cell(row=j, column=i).value
    #         fill = PatternFill('solid', fgColor=cell[1:])
    #         sheet.cell(row=j, column=i).fill = fill
    #         sheet.cell(row=j, column=i).value = ''
    # wb.save(result_path)
