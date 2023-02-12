import numpy as np
import os
from pandas.core.frame import DataFrame
import pandas as pd
from visdom import Visdom
import math
import random

from matplotlib import pyplot as plt
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import normalize
from sklearn.metrics import mean_squared_error, mean_absolute_error

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors

from pytorch_tabnet.tab_model import TabNetRegressor

from sklearn.metrics import r2_score
from sklearn.model_selection import train_test_split

import xgboost as xgb
import joblib
from pytorch_widedeep.models import TabResnet, WideDeep, TabMlp, TabTransformer
from pytorch_widedeep.preprocessing import TabPreprocessor
from pytorch_widedeep import Trainer
from pytorch_widedeep.initializers import KaimingNormal
import torch

column = 'mobile'  # kind mobile
row = 'xiaomi'  # kind : none green red water    mobile : xiaomi huawei redmi vivo


def ColourDistance(rgb_1, rgb_2):
    R_1, G_1, B_1 = rgb_1
    R_2, G_2, B_2 = rgb_2
    rmean = (R_1 + R_2) / 2
    R = R_1 - R_2
    G = G_1 - G_2
    B = B_1 - B_2
    return math.sqrt((2 + rmean / 256) * (R**2) + 4 * (G**2) + (2 + (255 - rmean) / 256) * (B**2))


def RGB_to_Hex(rgb):
    strs = '#'
    for i in rgb:
        num = int(float(i))  # 将str转int
        # 将R、G、B分别转化为16进制拼接转换并大写
        strs += str(hex(num))[-2:].replace('x', '0').upper()

    return strs


def arr2str(arr):
    # data = []
    hex_list = []
    for i in arr:
        # s = ','.join(str(j) for j in i)
        hex_list.append(RGB_to_Hex(i))
        # data.append(s)
    # data = np.array(data).reshape(-1, 1)
    hex_list = np.array(hex_list).reshape(-1, 1)
    return np.hstack((arr, hex_list))


def rmse(pred, label):
    tot = 0
    for i in range(len(pred)):
        rmse = 0
        for j in range(len(pred[i])):
            rmse += (pred[i][j] - label[i][j])**2
        rmse = np.sqrt(rmse / (len(pred[i])))
        tot += rmse

    return tot / (len(pred))


def rae(pred, label):
    rae = 0
    mae = 0
    label_aver = 0
    up = 0
    down = 0
    for i in range(len(pred)):
        for j in range(len(pred[i])):
            up += abs(pred[i][j] - label[i][j])
            down += label[i][j]
        mae += up
        label_aver += down
        # print(up, down, up / down)
    rae += up / down
    # print(mae / len(pred))

    return rae * 100


KIND = 'mobile'
VALUE = 'huawei'

if __name__ == '__main__':
    # 输入数据
    tot_path = r'D:\VScodeProjects\PythonProjects\ColorCorrection\Code\Dataset\tot_result.xlsx'
    # 输出表格
    result_path = r'D:\VScodeProjects\PythonProjects\ColorCorrection\Code\output_1222.xlsx'

    df = pd.read_excel(tot_path)
    # df = df.loc[df[KIND] == VALUE]
    df = df.loc[df['f_distance'] != 0.0]
    train_data, test_data = train_test_split(df, train_size=0.8, random_state=42, shuffle=True)
    data = np.array(test_data)
    data_input = np.hstack((data[:, 3:9], data[:, 10:12], data[:, 12:14]))
    data_label = data[:, 15:18]
    data_input = data_input.astype(np.float32)
    data_label = data_label.astype(np.float32)
    # print(data_input.shape)
    # print(data_label.shape)
    # input()

    # # 载入输入数据
    # df = pd.read_excel(train_Path)
    # train_data = np.array(df)
    # train_input = np.hstack((train_data[:, 3:9], train_data[:, 10:12], train_data[:, 12:14]))
    # train_input = train_input.astype(np.float32)

    # # 载入输出数据
    # # train_label = np.array(pd.read_csv(train_Path))
    # train_label = train_data[:, 14:17]
    # train_label = train_label.astype(np.float32)

    # cc = list(zip(train_input, train_label))
    # random.shuffle(cc)
    # train_input[:], train_label[:] = zip(*cc)

    # # 载入测试集
    # test_data = np.array(pd.read_excel(test_Path))
    # test_input = np.hstack((test_data[:, 3:9], test_data[:, 10:12], test_data[:, 12:14]))
    # test_input = test_input.astype(np.float32)

    # # test_label = pd.read_csv(test_Path)
    # test_label = test_data[:, 14:17]
    # test_label = test_label.astype(np.float32)

    # 网络
    net = TabNetRegressor()

    model_save_path = r'D:\VScodeProjects\PythonProjects\ColorCorrection\Code\Model_tabnet'
    net.load_model(model_save_path + '.zip')

    # 预测
    # train_pred = net.predict(train_input)
    # test_preds = net.predict(test_input)
    data_preds = net.predict(data_input)

    # 评价
    # print('train RMSE:{:.3f}'.format(np.sqrt(mean_squared_error(train_label, train_pred))))
    # print('train MAE:{:.3f}'.format(mean_absolute_error(train_label, train_pred)))
    # print('train R2:{:.3f}'.format(r2_score(train_pred, train_label)))
    # print('test RMSE:{:.3f}'.format(np.sqrt(mean_squared_error(test_preds, test_label))))
    # print('test MAE:{:.3F}'.format(mean_absolute_error(test_label, test_preds)))
    # print('test R2:{:.3f}'.format(r2_score(test_preds, test_label)))
    print('train myRMSE:{:.3f}'.format(rmse(data_preds, data_label)))
    print('data RMSE:{:.3f}'.format(np.sqrt(mean_squared_error(data_preds, data_label))))
    print('data MAE:{:.3F}'.format(mean_absolute_error(data_label, data_preds)))
    print('data R2:{:.3f}'.format(r2_score(data_preds, data_label)))
    print('tabnet:{:.3f}%'.format(rae(data_preds, data_label)))

    # 预测值保存
    preds = np.clip(data_preds, 0, 255).astype(int)
    preds = arr2str(preds)
    Input = arr2str(data_input[:, :3].astype(int))
    label = arr2str(data_label.astype(int))

    name = data[:, 2].reshape(-1, 1)
    mobile = data[:, 0].reshape(-1, 1)
    kind = data[:, 1].reshape(-1, 1)
    lightmode = data[:, 9].reshape(-1, 1)

    for i in range(len(lightmode)):
        if 'd50' in lightmode[i][0] or 'D50' in lightmode[i][0]:
            lightmode[i][0] = 'D50'
        elif 'd65' in lightmode[i][0] or 'D65' in lightmode[i][0]:
            lightmode[i][0] = 'D65'
        elif 'f' in lightmode[i][0] or 'F' in lightmode[i][0]:
            lightmode[i][0] = 'F'
        elif 'tl83' in lightmode[i][0] or 'Tl83' in lightmode[i][0]:
            lightmode[i][0] = 'Tl83'
        elif 'tl84' in lightmode[i][0] or 'Tl84' in lightmode[i][0]:
            lightmode[i][0] = 'Tl84'
        elif 'cloudy' in lightmode[i][0]:
            lightmode[i][0] = 'cloudy'
        elif 'max' in lightmode[i][0]:
            lightmode[i][0] = 'max'
        elif 'dark' in lightmode[i][0]:
            lightmode[i][0] = 'dark'
        elif 'light' in lightmode[i][0]:
            lightmode[i][0] = 'light'
        else:
            print(lightmode[i][0])
            input()

    # xgboost
    bst = xgb.Booster(model_file='XGB_model.xgb')  # 加载模型
    bst.load_model('./models/xgb/XGB_model.xgb')
    dtest = xgb.DMatrix(data_input)
    xgb_pred = bst.predict(dtest)
    print('xgboost:{:.3f}%'.format(rae(xgb_pred, data_label)))
    xgb_pred = np.clip(xgb_pred, 0, 255).astype(int)
    xgb_pred = arr2str(xgb_pred)

    # tree
    tree_model = joblib.load('./models/tree/tree.m')
    tree_pred = tree_model.predict(data_input)
    print('tree:{:.3f}%'.format(rae(tree_pred, data_label)))
    tree_pred = np.clip(tree_pred, 0, 255).astype(int)
    tree_pred = arr2str(tree_pred)

    # forest
    forest_model = joblib.load('./models/forest/forest.m')
    forest_pred = forest_model.predict(data_input)
    print('forest:{:.3f}%'.format(rae(forest_pred, data_label)))
    forest_pred = np.clip(forest_pred, 0, 255).astype(int)
    forest_pred = arr2str(forest_pred)

    # tabMlp
    continuous_cols = ["R", "G", "B", "R_P", "G_P", "R_P", "light", "exp", "iso", "f_distance"]
    tab_preprocessor = TabPreprocessor(continuous_cols=continuous_cols)
    test_input = tab_preprocessor.fit_transform(test_data)

    tabMlp = TabMlp(continuous_cols=continuous_cols,
                    mlp_batchnorm=True,
                    mlp_hidden_dims=[100, 5000, 5000, 100],
                    column_idx=tab_preprocessor.column_idx)
    tabResnet = TabResnet(continuous_cols=continuous_cols,
                          mlp_batchnorm=True,
                          blocks_dims=[200, 100, 100],
                          mlp_hidden_dims=[100, 5000, 5000, 100],
                          column_idx=tab_preprocessor.column_idx)
    tabTransformer = TabTransformer(continuous_cols=continuous_cols,
                                    embed_continuous=True,
                                    mlp_batchnorm=True,
                                    mlp_hidden_dims=[1000, 10000, 10000, 1000],
                                    shared_embed=True,
                                    column_idx=tab_preprocessor.column_idx)
    model = WideDeep(deeptabular=tabMlp, pred_dim=3)
    deep_opt = torch.optim.Adam(model.deeptabular.parameters())
    deep_sch = torch.optim.lr_scheduler.StepLR(deep_opt, step_size=3)
    optimizers = {"deeptabular": deep_opt}
    schedulers = {"deeptabular": deep_sch}
    initializers = {"deeptabular": KaimingNormal}
    model.load_state_dict(torch.load('./models/tabmlp/tabMlp.pt'))
    trainer = Trainer(model, objective='regression')
    tMlp_pred = trainer.predict(X_tab=test_input, batch_size=64)
    print('data RMSE:{:.3f}'.format(np.sqrt(mean_squared_error(tMlp_pred, data_label))))
    print('data MAE:{:.3F}'.format(mean_absolute_error(data_label, tMlp_pred)))
    print('data R2:{:.3f}'.format(r2_score(tMlp_pred, data_label)))
    print('tabMlp:{:.3f}%'.format(rae(tMlp_pred, data_label)))
    tMlp_pred = np.clip(tMlp_pred, 0, 255).astype(int)
    tMlp_pred = arr2str(tMlp_pred)

    # tabResnet
    model = WideDeep(deeptabular=tabResnet, pred_dim=3)
    model.load_state_dict(torch.load('./models/tabresnet/tabResnet.pt'))
    trainer = Trainer(model, objective='regression')
    tResnet_pred = trainer.predict(X_tab=test_input, batch_size=64)
    print('tabResnet:{:.3f}%'.format(rae(tResnet_pred, data_label)))
    tResnet_pred = np.clip(tResnet_pred, 0, 255).astype(int)
    tResnet_pred = arr2str(tResnet_pred)

    # tabTransformer
    model = WideDeep(deeptabular=tabTransformer, pred_dim=3)
    model.load_state_dict(torch.load('./models/tabtransformer/tabTransformer.pt'))
    trainer = Trainer(model, objective='regression')
    tTransformer_pred = trainer.predict(X_tab=test_input, batch_size=64)
    print('tabTransformer:{:.3f}%'.format(rae(tTransformer_pred, data_label)))
    tTransformer_pred = np.clip(tTransformer_pred, 0, 255).astype(int)
    tTransformer_pred = arr2str(tTransformer_pred)
    input()

    result = np.hstack((name, mobile, kind, lightmode, Input, label, preds, xgb_pred, tree_pred, forest_pred, tMlp_pred,
                        tResnet_pred, tTransformer_pred))

    dataframe = pd.DataFrame(result,
                             columns=[
                                 'name', 'mobile', 'kind', 'lightmode', 'input_R', 'input_G', 'input_B', 'input_color',
                                 'label_R', 'label_G', 'label_B', 'label_color', 'out_R', 'out_G', 'out_B', 'out_color',
                                 'out_R', 'out_G', 'out_B', 'out_color', 'out_R', 'out_G', 'out_B', 'out_color',
                                 'out_R', 'out_G', 'out_B', 'out_color', 'out_R', 'out_G', 'out_B', 'out_color',
                                 'out_R', 'out_G', 'out_B', 'out_color', 'out_R', 'out_G', 'out_B', 'out_color'
                             ])
    dataframe.to_excel(result_path, index=False)

    wb = load_workbook(result_path)
    sheet = wb.worksheets[0]

    for i in [8, 12, 16, 20, 24, 28, 32, 36, 40]:
        for j in range(2, len(data_input) + 2):
            cell = sheet.cell(row=j, column=i).value
            fill = PatternFill('solid', fgColor=cell[1:])
            sheet.cell(row=j, column=i).fill = fill
            sheet.cell(row=j, column=i).value = ''
    wb.save(result_path)
